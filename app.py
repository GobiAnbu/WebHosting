from flask import Flask, render_template, request, jsonify
import urllib.parse
import openpyxl
import math
import os
import glob
from datetime import datetime

app = Flask(__name__)

def get_chit_file():
    """Get the chit file name from query params or JSON body, default to chit.xlsx"""
    if request.method == "POST":
        data = request.get_json(silent=True)
        if data and data.get("chitFile"):
            return data.get("chitFile")
    return request.args.get("chitFile", "chit.xlsx")

@app.route("/")
def index():
    return render_template("home.html")

@app.route("/chit-details")
def chit_details():
    return render_template("ui.html")

@app.route("/reminder")
def reminder():
    return render_template("reminder.html")

@app.route("/chit-tomorrow")
def chit_tomorrow():
    return render_template("chit_tomorrow.html")

@app.route("/get-chit-files")
def get_chit_files():
    files = [os.path.basename(f) for f in glob.glob("*.xlsx") if not os.path.basename(f).startswith("~$")]
    return jsonify(files)

@app.route("/get-names")
def get_names():
    chit_file = get_chit_file()
    wb = openpyxl.load_workbook(chit_file)
    ws = wb["chitMembers"]
    headers = [cell.value for cell in ws[1]]
    name_idx = headers.index("Name")
    withdraw_idx = headers.index("Withdraw") if "Withdraw" in headers else None
    mobile_idx = headers.index("MobileNumber")
    names = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        withdraw = str(row[withdraw_idx]).strip().lower() if withdraw_idx is not None else "no"
        name = row[name_idx]
        if withdraw != "yes" and name and str(name).strip():
            names.append({"name": str(name).strip(), "mobile": str(row[mobile_idx])})
    wb.close()
    return jsonify(names)

@app.route("/get-chit-number")
def get_chit_number():
    chit_file = get_chit_file()
    wb = openpyxl.load_workbook(chit_file)
    ws = wb["chitNumberDetails"]
    now = datetime.now()
    chit_number = ""
    no_count = 0
    amount_per_person = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        month_val = row[2]  # Month column
        conducted = str(row[1]).strip().lower() if row[1] else ""
        if conducted == "no":
            no_count += 1
        if month_val and month_val.month == now.month and month_val.year == now.year:
            chit_number = row[0]  # Chit Number column
            amount_per_person = row[6] if len(row) > 6 and row[6] else 0
    wb.close()
    return jsonify({"chitNumber": chit_number, "chitRemaining": no_count - 1, "amountPerPerson": amount_per_person})

@app.route("/send-campaign", methods=["POST"])
def send_campaign():
    data = request.get_json()
    chit_file = data.get("chitFile", "chit.xlsx")

    # Update chitNumberDetails sheet
    chit_amount = data.get("chitAmount", "")
    discount_amount = data.get("discountAmount", "")
    name = data.get("name", "")
    amount_need_to_pay = math.ceil((float(chit_amount) / 20) / 10) * 10 if chit_amount else 0
    wb = openpyxl.load_workbook(chit_file)
    ws = wb["chitNumberDetails"]
    now = datetime.now()
    for row in ws.iter_rows(min_row=2):
        month_val = row[2].value  # Month column
        if month_val and month_val.month == now.month and month_val.year == now.year:
            row[1].value = "Yes"  # Conducted
            row[3].value = float(chit_amount) if chit_amount else None  # Chit Amount
            row[4].value = float(discount_amount) if discount_amount else None  # Discount Amount
            row[5].value = name  # Taken By
            row[6].value = amount_need_to_pay  # Amount Per Person
            break

    # Update Withdraw column in chitMembers for the selected name
    ws1 = wb["chitMembers"]
    headers1 = [cell.value for cell in ws1[1]]
    name_idx = headers1.index("Name")
    withdraw_idx = headers1.index("Withdraw")
    for row in ws1.iter_rows(min_row=2):
        if row[name_idx].value and str(row[name_idx].value).strip() == name.strip():
            row[withdraw_idx].value = "yes"
            break

    wb.save(chit_file)

    # Build WhatsApp message for all members in chitMembers
    ws_members = wb["chitMembers"]
    headers_m = [cell.value for cell in ws_members[1]]
    m_name_idx = headers_m.index("Name")
    m_mobile_idx = headers_m.index("MobileNumber")

    chit_number = data.get("chitNumber", "")
    total_amount = data.get("totalAmount", "")
    current_date = data.get("currentDate", "")
    chit_remaining = data.get("chitRemaining", "")

    # Format date to dd-mm-yyyy
    formatted_date = current_date
    if current_date:
        parts = current_date.split("-")
        if len(parts) == 3:
            formatted_date = f"{parts[2]}-{parts[1]}-{parts[0]}"

    msg = (
        f"Here are the chit details:\n\n"
        f"🎯 *Chit Taken By:* {name}\n"
        f"📌 *Chit Number:* {chit_number}\n"
        f"💰 *Total Amount:* ₹{total_amount}\n"
        f"🏷️ *Chit Amount:* ₹{chit_amount}\n"
        f"📅 *Date:* {formatted_date}\n"
        f"🔄 *Chit Remaining:* {chit_remaining}\n\n"
        f"💳 *Amount Need to Pay:* *₹{amount_need_to_pay}*\n\n"
        f"———————————\n\n"
        f"சிட் விவரங்கள்:\n\n"
        f"🎯 *சிட் எடுத்தவர்:* {name}\n"
        f"📌 *சிட் எண்:* {chit_number}\n"
        f"💰 *மொத்த தொகை:* ₹{total_amount}\n"
        f"🏷️ *சிட் தொகை:* ₹{chit_amount}\n"
        f"📅 *தேதி:* {formatted_date}\n"
        f"🔄 *மீதமுள்ள சிட்:* {chit_remaining}\n\n"
        f"💳 *செலுத்த வேண்டிய தொகை:* *₹{amount_need_to_pay}*\n\n"
        f"நன்றி! 🎉"
    )
    encoded = urllib.parse.quote(msg)

    results = []
    for row in ws_members.iter_rows(min_row=2, values_only=True):
        member_name = row[m_name_idx]
        mobile = row[m_mobile_idx]
        if member_name and str(member_name).strip() and mobile:
            phone = str(mobile).strip()
            wa_url = f"https://wa.me/91{phone}?text={encoded}"
            results.append({"name": str(member_name).strip(), "phone": phone, "url": wa_url})

    wb.close()
    return jsonify({"success": True, "results": results})

@app.route("/send-reminder", methods=["POST"])
def send_reminder():
    data = request.get_json()
    message = data.get("message", "")
    members = data.get("members", [])
    chit_file = data.get("chitFile", "chit.xlsx")

    # Get amount per person from chitNumberDetails
    wb = openpyxl.load_workbook(chit_file)
    ws = wb["chitNumberDetails"]
    now = datetime.now()
    amount_per_person = "0"
    for row in ws.iter_rows(min_row=2, values_only=True):
        month_val = row[2]
        if month_val and month_val.month == now.month and month_val.year == now.year:
            amount_per_person = str(row[6]) if len(row) > 6 and row[6] else "0"
            break
    wb.close()

    results = []
    for member in members:
        name = member.get("name", "")
        phone = str(member.get("mobile", "")).strip()
        msg = message.replace("{name}", name).replace("{amount}", amount_per_person)
        encoded = urllib.parse.quote(msg)
        wa_url = f"https://wa.me/91{phone}?text={encoded}"
        results.append({"name": name, "phone": phone, "url": wa_url})
    return jsonify({"success": True, "results": results})

@app.route("/send-tomorrow-reminder", methods=["POST"])
def send_tomorrow_reminder():
    data = request.get_json()
    message = data.get("message", "")
    chit_date = data.get("date", "")
    chit_number = data.get("chitNumber", "")
    chit_file = data.get("chitFile", "chit.xlsx")

    # Get all members from chitMembers sheet
    wb = openpyxl.load_workbook(chit_file)
    ws = wb["chitMembers"]
    headers = [cell.value for cell in ws[1]]
    name_idx = headers.index("Name")
    mobile_idx = headers.index("MobileNumber")

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        member_name = row[name_idx]
        mobile = row[mobile_idx]
        if member_name and str(member_name).strip() and mobile:
            name = str(member_name).strip()
            phone = str(mobile).strip()
            msg = message.replace("{name}", name).replace("{date}", chit_date).replace("{chitNumber}", chit_number)
            encoded = urllib.parse.quote(msg)
            wa_url = f"https://wa.me/91{phone}?text={encoded}"
            results.append({"name": name, "phone": phone, "url": wa_url})
    wb.close()
    return jsonify({"success": True, "results": results})

if __name__ == "__main__":
    app.run(debug=True, port=5001)
